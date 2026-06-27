from __future__ import annotations

import csv
import io
import json

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Query, status
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import require_role
from ..minio_client import upload_bytes
from ..models import Dataset, DatasetFile, DatasetItem, User
from ..response import ok
from ..schemas import DatasetCreate, MapFieldsRequest

router = APIRouter(prefix="/api/datasets", tags=["datasets"])


def _out(d: Dataset) -> dict:
    return {
        "id": d.id,
        "name": d.name,
        "description": d.description,
        "format": d.format,
        "file_count": d.file_count,
        "item_count": d.item_count,
        "status": d.status,
        "field_mapping": d.field_mapping_json,
        "created_by": d.created_by,
        "created_at": d.created_at,
    }


def _detect_format(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".jsonl"):
        return "jsonl"
    if lower.endswith(".json"):
        return "json"
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    return "json"


def parse_rows(data: bytes, fmt: str) -> list[dict]:
    if fmt == "json":
        text = data.decode("utf-8-sig")
        obj = json.loads(text)
        if isinstance(obj, list):
            return [r if isinstance(r, dict) else {"value": r} for r in obj]
        if isinstance(obj, dict):
            return [obj]
        return []
    if fmt == "jsonl":
        rows = []
        for line in data.decode("utf-8-sig").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if isinstance(obj, dict):
                rows.append(obj)
            else:
                rows.append({"value": obj})
        return rows
    if fmt == "csv":
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]
    if fmt == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if headers is None:
            return []
        keys = [str(h) if h is not None else f"col{i}" for i, h in enumerate(headers)]
        out: list[dict] = []
        for row in rows_iter:
            if row is None:
                continue
            if all(v is None for v in row):
                continue
            out.append({keys[i]: (row[i] if i < len(row) else None) for i in range(len(keys))})
        return out
    return []


@router.get("")
def list_datasets(db: Session = Depends(get_db), _user: User = Depends(require_role("owner", "admin", "labeler", "reviewer"))):
    rows = db.query(Dataset).order_by(Dataset.id.desc()).all()
    return ok([_out(d) for d in rows])


@router.post("")
def create_dataset(body: DatasetCreate, db: Session = Depends(get_db), user: User = Depends(require_role("owner", "admin"))):
    d = Dataset(name=body.name, description=body.description, created_by=user.id)
    db.add(d)
    db.commit()
    db.refresh(d)
    return ok(_out(d))


@router.get("/{dataset_id}")
def get_dataset(dataset_id: int, db: Session = Depends(get_db), _user: User = Depends(require_role("owner", "admin", "labeler", "reviewer"))):
    d = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if d is None:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "dataset not found", "data": None})
    return ok(_out(d))


@router.delete("/{dataset_id}")
def delete_dataset(dataset_id: int, db: Session = Depends(get_db), _user: User = Depends(require_role("owner", "admin"))):
    d = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if d is None:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "dataset not found", "data": None})
    db.query(DatasetItem).filter(DatasetItem.dataset_id == dataset_id).delete()
    db.query(DatasetFile).filter(DatasetFile.dataset_id == dataset_id).delete()
    db.delete(d)
    db.commit()
    return ok({})


@router.post("/{dataset_id}/upload")
def upload_dataset(
    dataset_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _user: User = Depends(require_role("owner", "admin")),
):
    d = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if d is None:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "dataset not found", "data": None})
    content = file.file.read()
    fmt = _detect_format(file.filename or "data.json")
    try:
        rows = parse_rows(content, fmt)
    except Exception as e:
        raise HTTPException(status_code=400, detail={"code": 400, "message": f"parse error: {e}", "data": None})

    object_name = upload_bytes("dataset", content, file.filename or "data", "application/octet-stream")
    df = DatasetFile(
        dataset_id=d.id,
        filename=file.filename or "data",
        minio_object=object_name,
        size=len(content),
        format=fmt,
        status="uploaded",
    )
    db.add(df)
    db.flush()

    start = db.query(DatasetItem).filter(DatasetItem.dataset_id == d.id).count()
    for i, row in enumerate(rows):
        db.add(DatasetItem(dataset_id=d.id, file_id=df.id, index=start + i, raw_data=row, status="pending"))

    d.format = fmt
    d.file_count = (d.file_count or 0) + 1
    d.item_count = (d.item_count or 0) + len(rows)
    db.commit()
    db.refresh(d)
    return ok(_out(d))


@router.get("/{dataset_id}/items")
def list_items(
    dataset_id: int,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=500),
    db: Session = Depends(get_db),
    _user: User = Depends(require_role("owner", "admin", "labeler", "reviewer")),
):
    d = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if d is None:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "dataset not found", "data": None})
    total = db.query(DatasetItem).filter(DatasetItem.dataset_id == dataset_id).count()
    rows = (
        db.query(DatasetItem)
        .filter(DatasetItem.dataset_id == dataset_id)
        .order_by(DatasetItem.index)
        .offset((page - 1) * size)
        .limit(size)
        .all()
    )
    fields = set()
    for r in rows:
        if isinstance(r.raw_data, dict):
            fields.update(r.raw_data.keys())
    return ok({
        "items": [
            {
                "id": r.id,
                "dataset_id": r.dataset_id,
                "index": r.index,
                "raw_data": r.raw_data,
                "mapped_fields": r.mapped_fields,
                "status": r.status,
            }
            for r in rows
        ],
        "total": total,
        "fields": sorted(fields),
    })


@router.post("/{dataset_id}/map-fields")
def map_fields(
    dataset_id: int,
    body: MapFieldsRequest,
    db: Session = Depends(get_db),
    _user: User = Depends(require_role("owner", "admin")),
):
    d = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if d is None:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "dataset not found", "data": None})
    d.field_mapping_json = body.mapping
    items = db.query(DatasetItem).filter(DatasetItem.dataset_id == dataset_id).all()
    for it in items:
        mapped = {}
        raw = it.raw_data or {}
        for tkey, dfield in body.mapping.items():
            mapped[tkey] = raw.get(dfield) if isinstance(raw, dict) else None
        it.mapped_fields = mapped
    db.commit()
    db.refresh(d)
    return ok(_out(d))